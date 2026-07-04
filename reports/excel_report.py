"""
Excel Report Generator using openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
from typing import Dict, Any, List, Optional, Union


def add_styled_header(ws: Any, headers: List[str]) -> None:
    """
    Add styled header row to worksheet.
    
    Args:
        ws: openpyxl worksheet
        headers: List of header strings
    """
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Set row height
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'


def add_styled_data(ws: Any, data: List[List[Any]], start_row: int = 2) -> None:
    """
    Add styled data rows to worksheet.
    
    Args:
        ws: openpyxl worksheet
        data: List of data rows
        start_row: Starting row number
    """
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for row_idx, row_data in enumerate(data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            
            # Alternate row coloring
            if (row_idx - start_row) % 2 == 1:
                cell.fill = alt_fill


def add_charts(ws: Any, data: List[List[Any]], chart_type: str = 'bar',
               data_start_row: int = 2, category_col: int = 1, value_col: int = 2) -> None:
    """
    Add charts to worksheet.
    
    Args:
        ws: openpyxl worksheet
        data: Data for the chart
        chart_type: Type of chart ('bar', 'line', 'pie')
        data_start_row: Starting row of data
        category_col: Column index for categories (1-based)
        value_col: Column index for values (1-based)
    """
    data_end_row = data_start_row + len(data) - 1
    
    if chart_type == 'bar':
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Data Analysis"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Category"
    elif chart_type == 'line':
        chart = LineChart()
        chart.style = 10
        chart.title = "Trend Analysis"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Period"
    else:  # pie
        chart = BarChart()
        chart.type = "pie"
        chart.style = 10
        chart.title = "Distribution"
    
    # Add data
    values = Reference(ws, min_col=value_col, min_row=data_start_row - 1, 
                      max_row=data_end_row, max_col=value_col)
    categories = Reference(ws, min_col=category_col, min_row=data_start_row, 
                          max_row=data_end_row)
    
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    
    # Position chart
    chart.width = 20
    chart.height = 12
    
    # Add chart to worksheet
    ws.add_chart(chart, f"H{data_start_row}")


def create_property_excel(property_data: Dict[str, Any], output_path: str) -> str:
    """
    Create Excel report for a single property.
    
    Args:
        property_data: Dictionary containing property information
        output_path: Path to save the Excel file
        
    Returns:
        Path to saved Excel file
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Property Details"
    
    # Property information section
    ws.cell(row=1, column=1, value="PROPERTY INFORMATION").font = Font(bold=True, size=14, color='1F4E79')
    ws.merge_cells('A1:D1')
    
    # Headers
    headers = ['Field', 'Value']
    add_styled_header(ws, headers)
    
    # Property data
    field_mapping = {
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'zip_code': 'ZIP Code',
        'property_type': 'Property Type',
        'bedrooms': 'Bedrooms',
        'bathrooms': 'Bathrooms',
        'area_sqft': 'Area (sqft)',
        'year_built': 'Year Built',
        'condition': 'Condition',
        'price': 'Listing Price',
        'lot_size': 'Lot Size',
        'garage_spaces': 'Garage Spaces',
        'description': 'Description'
    }
    
    data_rows = []
    for field, label in field_mapping.items():
        if field in property_data and property_data[field] is not None:
            value = property_data[field]
            if field == 'price':
                value = f"${value:,.2f}"
            elif field == 'area_sqft':
                value = f"{value:,.0f} sqft"
            data_rows.append([label, str(value)])
    
    add_styled_data(ws, data_rows)
    
    # Auto-fit columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    
    # Financial metrics section if available
    if any(key in property_data for key in ['rental_yield', 'cap_rate', 'investment_score']):
        start_row = len(data_rows) + 4
        ws.cell(row=start_row, column=1, value="FINANCIAL METRICS").font = Font(bold=True, size=14, color='27AE60')
        ws.merge_cells(f'A{start_row}:D{start_row}')
        
        financial_data = []
        if 'rental_yield' in property_data:
            financial_data.append(['Rental Yield', f"{property_data['rental_yield']:.2f}%"])
        if 'cap_rate' in property_data:
            financial_data.append(['Cap Rate', f"{property_data['cap_rate']:.2f}%"])
        if 'investment_score' in property_data:
            financial_data.append(['Investment Score', f"{property_data['investment_score']:.1f}/100"])
        if 'annual_rent' in property_data:
            financial_data.append(['Annual Rent', f"${property_data['annual_rent']:,.2f}"])
        
        add_styled_header(ws, headers)
        add_styled_data(ws, financial_data, start_row=start_row + 1)
    
    wb.save(output_path)
    return output_path


def create_market_excel(df: pd.DataFrame, output_path: str) -> str:
    """
    Create Excel report for market analysis.
    
    Args:
        df: Input DataFrame with market data
        output_path: Path to save the Excel file
        
    Returns:
        Path to saved Excel file
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Market Overview"
    
    ws_overview.cell(row=1, column=1, value="MARKET OVERVIEW").font = Font(bold=True, size=14, color='1F4E79')
    ws_overview.merge_cells('A1:F1')
    
    # Summary statistics
    if 'price' in df.columns:
        stats = [
            ['Total Listings', len(df)],
            ['Average Price', f"${df['price'].mean():,.2f}"],
            ['Median Price', f"${df['price'].median():,.2f}"],
            ['Price Std Dev', f"${df['price'].std():,.2f}"],
            ['Min Price', f"${df['price'].min():,.2f}"],
            ['Max Price', f"${df['price'].max():,.2f}"]
        ]
        
        add_styled_header(ws_overview, ['Metric', 'Value'])
        add_styled_data(ws_overview, stats)
        
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 25
    
    # Sheet 2: By City
    if 'city' in df.columns and 'price' in df.columns:
        ws_city = wb.create_sheet("By City")
        
        ws_city.cell(row=1, column=1, value="ANALYSIS BY CITY").font = Font(bold=True, size=14, color='8E44AD')
        ws_city.merge_cells('A1:F1')
        
        city_stats = df.groupby('city').agg({
            'price': ['mean', 'median', 'count'],
            'area_sqft': 'mean' if 'area_sqft' in df.columns else lambda x: None
        }).round(2)
        
        city_stats.columns = ['Avg Price', 'Median Price', 'Listing Count', 'Avg Area']
        city_stats = city_stats.reset_index()
        
        headers = ['City'] + list(city_stats.columns[1:])
        add_styled_header(ws_city, headers)
        
        data_rows = city_stats.values.tolist()
        add_styled_data(ws_city, data_rows)
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws_city.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # Add chart
        add_charts(ws_city, data_rows, 'bar', start_row=3, category_col=1, value_col=2)
    
    # Sheet 3: By Property Type
    if 'property_type' in df.columns and 'price' in df.columns:
        ws_type = wb.create_sheet("By Property Type")
        
        ws_type.cell(row=1, column=1, value="ANALYSIS BY PROPERTY TYPE").font = Font(bold=True, size=14, color='E74C3C')
        ws_type.merge_cells('A1:F1')
        
        type_stats = df.groupby('property_type').agg({
            'price': ['mean', 'median', 'count'],
            'bedrooms': 'mean' if 'bedrooms' in df.columns else lambda x: None
        }).round(2)
        
        type_stats.columns = ['Avg Price', 'Median Price', 'Listing Count', 'Avg Bedrooms']
        type_stats = type_stats.reset_index()
        
        headers = ['Property Type'] + list(type_stats.columns[1:])
        add_styled_header(ws_type, headers)
        
        data_rows = type_stats.values.tolist()
        add_styled_data(ws_type, data_rows)
        
        for col in range(1, len(headers) + 1):
            ws_type.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        add_charts(ws_type, data_rows, 'bar', start_row=3, category_col=1, value_col=2)
    
    # Sheet 4: Raw Data
    ws_raw = wb.create_sheet("Raw Data")
    
    ws_raw.cell(row=1, column=1, value="RAW DATA EXPORT").font = Font(bold=True, size=14, color='2C3E50')
    ws_raw.merge_cells('A1:F1')
    
    # Write DataFrame to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_raw.append(r)
    
    # Style header
    for cell in ws_raw[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    # Auto-fit columns (approximate)
    for col in range(1, len(df.columns) + 1):
        ws_raw.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    wb.save(output_path)
    return output_path


def create_comparison_excel(properties_list: List[Dict[str, Any]], output_path: str) -> str:
    """
    Create Excel report for comparing multiple properties.
    
    Args:
        properties_list: List of property dictionaries
        output_path: Path to save the Excel file
        
    Returns:
        Path to saved Excel file
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Property Comparison"
    
    ws.cell(row=1, column=1, value="PROPERTY COMPARISON").font = Font(bold=True, size=14, color='1F4E79')
    ws.merge_cells('A1:J1')
    
    # Define comparison fields
    comparison_fields = [
        'address', 'city', 'price', 'bedrooms', 'bathrooms', 
        'area_sqft', 'year_built', 'property_type', 'condition',
        'rental_yield', 'investment_score'
    ]
    
    headers = ['Property'] + [field.replace('_', ' ').title() for field in comparison_fields]
    add_styled_header(ws, headers)
    
    # Add property data
    data_rows = []
    for i, prop in enumerate(properties_list, 1):
        row = [f"Property {i}"]
        for field in comparison_fields:
            value = prop.get(field, '-')
            if field == 'price' and isinstance(value, (int, float)):
                value = f"${value:,.2f}"
            elif field == 'area_sqft' and isinstance(value, (int, float)):
                value = f"{value:,.0f}"
            elif field in ['rental_yield', 'investment_score'] and isinstance(value, (int, float)):
                value = f"{value:.1f}"
            row.append(str(value) if value else '-')
        data_rows.append(row)
    
    add_styled_data(ws, data_rows)
    
    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Add comparison chart for price
    if any('price' in prop for prop in properties_list):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Price Comparison"
        chart.y_axis.title = "Price ($)"
        
        values = Reference(ws, min_col=3, min_row=2, max_row=len(properties_list) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(properties_list) + 1)
        
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 20
        chart.height = 12
        
        ws.add_chart(chart, "A" + str(len(properties_list) + 4))
    
    wb.save(output_path)
    return output_path


def create_data_export(df: pd.DataFrame, output_path: str) -> str:
    """
    Export raw data to Excel with basic formatting.
    
    Args:
        df: Input DataFrame
        output_path: Path to save the Excel file
        
    Returns:
        Path to saved Excel file
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Export"
    
    # Write header
    headers = list(df.columns)
    add_styled_header(ws, headers)
    
    # Write data
    data_rows = df.values.tolist()
    add_styled_data(ws, data_rows)
    
    # Auto-fit columns (approximate)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Add filters
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(data_rows) + 1}"
    
    wb.save(output_path)
    return output_path
