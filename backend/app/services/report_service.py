import io
import csv
from typing import Any, Dict, List
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportService:
    def generate_pdf(self, property_data: Dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=24, spaceAfter=20, textColor=colors.HexColor("#1a365d"))
        heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=16, spaceAfter=10, textColor=colors.HexColor("#2d3748"))
        body_style = ParagraphStyle("Body", parent=styles["Normal"], fontSize=11, spaceAfter=8, textColor=colors.HexColor("#4a5568"))

        elements.append(Paragraph(property_data.get("title", "Property Report"), title_style))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Property Details", heading_style))

        details = [
            ["Property Type", property_data.get("property_type", "N/A")],
            ["Price", f"${property_data.get('price', 0):,.2f}"],
            ["City", property_data.get("city", "N/A")],
            ["Locality", property_data.get("locality", "N/A")],
            ["Address", property_data.get("address", "N/A")],
            ["Bedrooms", str(property_data.get("bedrooms", "N/A"))],
            ["Bathrooms", str(property_data.get("bathrooms", "N/A"))],
            ["Area", f"{property_data.get('area', 0)} sqft"],
            ["Parking", str(property_data.get("parking", 0))],
            ["Furnished", "Yes" if property_data.get("furnished") else "No"],
            ["Year Built", str(property_data.get("year_built", "N/A"))],
            ["Property Age", str(property_data.get("property_age", "N/A"))],
        ]

        table = Table(details, colWidths=[2 * inch, 4 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#edf2f7")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2d3748")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("Location Analysis", heading_style))
        location_data = [
            ["Hospital Distance", f"{property_data.get('hospital_distance', 'N/A')} km"],
            ["Metro Distance", f"{property_data.get('metro_distance', 'N/A')} km"],
            ["Crime Rate", str(property_data.get("crime_rate", "N/A"))],
            ["Population", str(property_data.get("population", "N/A"))],
        ]
        loc_table = Table(location_data, colWidths=[2 * inch, 4 * inch])
        loc_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#edf2f7")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2d3748")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ]))
        elements.append(loc_table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("Investment Analysis", heading_style))
        invest_data = [
            ["Rental Yield", f"{property_data.get('rental_yield', 0):.2f}%"],
            ["Market Growth", f"{property_data.get('market_growth', 0):.2f}%"],
        ]
        inv_table = Table(invest_data, colWidths=[2 * inch, 4 * inch])
        inv_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#edf2f7")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2d3748")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ]))
        elements.append(inv_table)
        elements.append(Spacer(1, 20))

        if property_data.get("description"):
            elements.append(Paragraph("Description", heading_style))
            elements.append(Paragraph(property_data["description"], body_style))

        if property_data.get("amenities"):
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("Amenities", heading_style))
            amenities_text = ", ".join(property_data["amenities"]) if isinstance(property_data["amenities"], list) else str(property_data["amenities"])
            elements.append(Paragraph(amenities_text, body_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    def generate_excel(self, data: List[Dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Properties"

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if data:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, record in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            for col_idx, header in enumerate(headers, 1):
                max_length = max(len(str(header)), max((len(str(record.get(header, ""))) for record in data), default=0))
                ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_length + 2, 30)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_csv(self, data: List[Dict[str, Any]]) -> str:
        if not data:
            return ""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()
